from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from fastapi import UploadFile
from sqlalchemy.orm import Session
import io
import os
import uuid
from PIL import Image

from app.models import Box, Focal, Voter


def extract_images_from_excel(file_content: bytes) -> Dict[int, str]:
    """
    Extract embedded images from Excel file and save them to uploads folder.
    Returns a dictionary mapping row numbers to saved image paths.
    """
    row_images = {}

    try:
        wb = load_workbook(filename=io.BytesIO(file_content))
        ws = wb.active

        # Ensure uploads directory exists
        uploads_dir = "uploads"
        os.makedirs(uploads_dir, exist_ok=True)

        # Get all images from the worksheet
        if hasattr(ws, '_images'):
            for img in ws._images:
                try:
                    # Get the anchor position (row and column)
                    if hasattr(img, 'anchor'):
                        anchor = img.anchor
                        if hasattr(anchor, '_from'):
                            row = anchor._from.row + 1  # Convert to 1-indexed
                        elif hasattr(anchor, 'row'):
                            row = anchor.row + 1
                        else:
                            continue
                    else:
                        continue

                    # Extract image data
                    if hasattr(img, '_data'):
                        img_data = img._data()
                    elif hasattr(img, 'ref'):
                        img_data = img.ref
                    else:
                        continue

                    # Generate unique filename
                    filename = f"{uuid.uuid4().hex}.png"
                    filepath = os.path.join(uploads_dir, filename)

                    # Save image
                    if isinstance(img_data, bytes):
                        pil_img = Image.open(io.BytesIO(img_data))
                        pil_img.save(filepath, "PNG")
                        row_images[row] = filename
                except Exception:
                    continue

        wb.close()
    except Exception:
        pass

    return row_images


def parse_excel(file_content: bytes, extract_photos: bool = False) -> Tuple[List[Dict[str, Any]], List[str], Dict[int, str]]:
    """
    Parse Excel file and return rows as list of dictionaries.
    Returns (data_rows, errors, row_images)
    """
    errors = []
    rows = []
    row_images = {}

    # Extract images if requested
    if extract_photos:
        row_images = extract_images_from_excel(file_content)

    try:
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
        ws = wb.active

        # Get header row (first row)
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        # Find photo column index if it exists
        photo_col_idx = None
        for idx, header in enumerate(headers):
            if header.lower() in ('photo', 'image', 'picture', 'pic'):
                photo_col_idx = idx
                break

        # Process data rows
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {}
            row_data['_row_num'] = row_num  # Store row number for image matching

            for idx, value in enumerate(row):
                if idx < len(headers):
                    header = headers[idx]
                    row_data[header] = value

            # Skip empty rows
            name = row_data.get("Name") or row_data.get("name")
            if name:
                rows.append(row_data)

        wb.close()

    except Exception as e:
        errors.append(f"Error parsing Excel file: {str(e)}")

    return rows, errors, row_images


def import_voters_from_excel(
    db: Session,
    file_content: bytes,
    import_photos: bool = True
) -> Dict[str, Any]:
    """
    Import voters from Excel file.
    Auto-creates boxes and focals.
    Optionally imports embedded photos.
    Returns import statistics.
    """
    stats = {
        "total_rows": 0,
        "imported": 0,
        "skipped": 0,
        "boxes_created": 0,
        "focals_created": 0,
        "photos_imported": 0,
        "errors": []
    }

    rows, parse_errors, row_images = parse_excel(file_content, extract_photos=import_photos)
    stats["errors"].extend(parse_errors)
    stats["total_rows"] = len(rows)

    if parse_errors:
        return stats

    # Cache for boxes and focals
    box_cache: Dict[str, Box] = {}
    focal_cache: Dict[str, Focal] = {}

    # Pre-load existing boxes and focals
    for box in db.query(Box).all():
        box_cache[box.name.lower()] = box

    for focal in db.query(Focal).all():
        focal_cache[focal.name.lower()] = focal

    for row_num, row in enumerate(rows, start=2):
        try:
            # Extract voter data
            name = row.get("Name") or row.get("name")
            if not name:
                stats["skipped"] += 1
                continue

            # EC # - Election Commission number
            ec_number = None
            ec_val = row.get("EC #") or row.get("EC#")
            if ec_val:
                try:
                    ec_number = int(ec_val)
                except (ValueError, TypeError):
                    pass

            # # column -> voter_id (sequence number)
            voter_id = str(row.get("#") or "").strip() or None

            # ID column -> national_id (ID card number like A359937)
            national_id = str(row.get("ID") or row.get("id") or "").strip() or None

            # Check for duplicate national_id
            if national_id:
                existing = db.query(Voter).filter(Voter.national_id == national_id).first()
                if existing:
                    stats["skipped"] += 1
                    stats["errors"].append(f"Row {row_num}: Duplicate national ID '{national_id}'")
                    continue

            # Get or create box
            box = None
            box_name = row.get("Registered Box") or row.get("Box#") or row.get("box")
            if box_name:
                box_name_str = str(box_name).strip()
                box_key = box_name_str.lower()
                if box_key not in box_cache:
                    new_box = Box(name=box_name_str)
                    db.add(new_box)
                    db.flush()
                    box_cache[box_key] = new_box
                    stats["boxes_created"] += 1
                box = box_cache[box_key]

            # Parse focals (comma-separated)
            voter_focals = []
            focal_str = row.get("Focal") or row.get("focal") or ""
            if focal_str:
                focal_names = [f.strip() for f in str(focal_str).split(",") if f.strip()]
                for focal_name in focal_names:
                    focal_key = focal_name.lower()
                    if focal_key not in focal_cache:
                        new_focal = Focal(name=focal_name)
                        db.add(new_focal)
                        db.flush()
                        focal_cache[focal_key] = new_focal
                        stats["focals_created"] += 1
                    voter_focals.append(focal_cache[focal_key])

            # Determine pledged status
            is_pledged = False
            pledged_val = row.get("Pledged") or row.get("Y") or row.get("pledged")
            if pledged_val:
                pledged_str = str(pledged_val).upper().strip()
                is_pledged = pledged_str in ("Y", "YES", "TRUE", "1")

            # Check for photo from embedded images
            photo_path = None
            original_row_num = row.get('_row_num')
            if original_row_num and original_row_num in row_images:
                photo_path = row_images[original_row_num]
                stats["photos_imported"] += 1

            # Box# column (e.g., "B2.1")
            box_number = str(row.get("Box#") or "").strip() or None

            # Remarks
            remarks = str(row.get("Remarks") or row.get("remarks") or "").strip() or None

            # Create voter
            voter = Voter(
                ec_number=ec_number,
                voter_id=voter_id,
                national_id=national_id,
                name=str(name).strip(),
                gender=str(row.get("G") or row.get("Gender") or row.get("gender") or "").strip() or None,
                age=int(row.get("Age") or row.get("age") or 0) if row.get("Age") or row.get("age") else None,
                party=str(row.get("P") or row.get("Party") or row.get("party") or "").strip() or None,
                address=str(row.get("Address") or row.get("address") or "").strip() or None,
                contact=str(row.get("Contact") or row.get("contact") or "").strip() or None,
                new_contact=str(row.get("New Contact") or "").strip() or None,
                previous_island=str(row.get("Previous Island") or "").strip() or None,
                previous_address=str(row.get("Previous address") or "").strip() or None,
                current_location=str(row.get("Current Location") or "").strip() or None,
                box_number=box_number,
                zone=str(row.get("Zone") or row.get("zone") or "").strip() or None,
                focal_comment=str(row.get("Focal Comment") or "").strip() or None,
                remarks=remarks,
                box=box,
                is_pledged=is_pledged,
                focals=voter_focals,
                photo_path=photo_path
            )

            db.add(voter)
            stats["imported"] += 1

        except Exception as e:
            stats["errors"].append(f"Row {row_num}: {str(e)}")
            stats["skipped"] += 1

    db.commit()
    return stats
