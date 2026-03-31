import io
import os

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from starlette.responses import StreamingResponse

from app.database import get_db
from app.models import User, Voter
from app.models.user import UserRole
from app.services.auth import get_current_user_required, require_role
from app.services.excel_import import import_voters_from_excel
from app.services.logging import log_activity, Actions

router = APIRouter(prefix="/import", tags=["Import"])
templates = Jinja2Templates(directory="app/templates")


@router.get("", response_class=HTMLResponse)
async def import_page(
    request: Request,
    user: User = Depends(require_role(UserRole.admin))
):
    """Render the import page (admin only)."""
    return templates.TemplateResponse(
        "import.html",
        {"request": request, "user": user}
    )


@router.get("/sample")
async def download_sample(
    user: User = Depends(get_current_user_required)
):
    """Download sample Excel file."""
    sample_path = os.path.join("Sample Data", "Sample Data.xlsx")
    if not os.path.exists(sample_path):
        raise HTTPException(status_code=404, detail="Sample file not found")
    return FileResponse(
        sample_path,
        filename="VoteCouncil_Sample_Data.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/export-current")
async def export_current_data(
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.admin))
):
    """Export all current voter data as Excel file."""
    from sqlalchemy.orm import joinedload
    voters = db.query(Voter).options(
        joinedload(Voter.box),
        joinedload(Voter.focals)
    ).order_by(Voter.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Voters"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")

    headers = [
        "EC#", "#", "ID", "Name", "Gender", "Age", "Party",
        "Address", "Contact", "Current Location", "Registered Box", "Box#",
        "Zone", "Focal", "Focal Comment", "Remarks", "Pledged", "Vote Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, v in enumerate(voters, 2):
        focals = ", ".join(f.name for f in v.focals) if v.focals else ""
        pledge = "Yes" if v.is_pledged else "No"
        status = v.vote_status.value.replace("_", " ").title() if v.vote_status else ""

        values = [
            v.ec_number, v.voter_id, v.national_id, v.name,
            v.gender, v.age, v.party, v.address,
            v.contact, v.current_location,
            v.box.name if v.box else "", v.box_number,
            v.zone, focals, v.focal_comment, v.remarks,
            pledge, status
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=VoteCouncil_Current_Data.xlsx"}
    )


@router.post("/excel")
async def import_excel(
    request: Request,
    file: UploadFile = File(...),
    import_photos: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.admin))
):
    """
    Import voters from Excel file (admin only).
    Auto-creates boxes and focals as needed.
    Optionally imports embedded photos.
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload an Excel file (.xlsx or .xls)"
        )

    # Read file content
    content = await file.read()

    # Import voters
    stats = import_voters_from_excel(db, content, import_photos=import_photos)

    log_activity(db, Actions.IMPORT_COMPLETE, user=user, details=f"Excel import from '{file.filename}': {stats.get('created', 0)} created, {stats.get('updated', 0)} updated", entity_type="Import", ip_address=request.client.host if request.client else None)

    return stats
